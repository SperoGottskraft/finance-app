"""
seed_all_data.py
================
One-time import script — run from the backend/ directory:
    python seed_all_data.py

Imports all CSV/XLSX data files into the finance database:
  • Checking.csv               → accounts + transactions
  • Savings.csv                → accounts + transactions
  • USAA CreditCard.csv        → accounts + transactions
  • ByBenefitType_expanded.xlsx → investment_accounts + investment_holdings (Fidelity 401K)
  • FidelityTotalValue.xlsx    → investment_accounts (Fidelity total snapshot)
  • charlesschwab copy.xlsx    → investment_accounts + investment_holdings
  • Edward Jones holdings_*.csv → investment_accounts + investment_holdings
  • ETradePortfolioDownload.csv → investment_accounts + investment_holdings
  • Irregular Expense History.csv → irregular_expenses
  • Regular Expense History.csv   → regular_expenses
"""

import os
import sys
import re
import json
from datetime import datetime, date, timezone
from pathlib import Path

import pandas as pd
from dateutil import parser as dateparser
from sqlalchemy.orm import Session

# ── Path setup ────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent          # backend/
DATA_DIR   = SCRIPT_DIR.parent              # finance_app/
sys.path.insert(0, str(SCRIPT_DIR))

from app.database import SessionLocal, engine, Base
from app.models import (
    Account, Transaction, Category,
    InvestmentAccount, InvestmentHolding,
    IrregularExpense, RegularExpense,
)

Base.metadata.create_all(bind=engine)

# ── Helpers ───────────────────────────────────────────────────────────────────

def to_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val) if not pd.isna(val) else 0.0
    s = str(val).strip().replace(",", "").replace("$", "").replace("%", "")
    s = re.sub(r"\((.+?)\)", r"-\1", s)   # (123.45) → -123.45
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def parse_date_utc(val) -> datetime | None:
    if not val or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        if isinstance(val, (datetime, pd.Timestamp)):
            dt = pd.Timestamp(val).to_pydatetime()
        else:
            dt = dateparser.parse(str(val))
        if dt is None:
            return None
        return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt
    except Exception:
        return None


def parse_date_only(val):
    """Return a Python date (no tz)."""
    dt = parse_date_utc(val)
    return dt.date() if dt else None


def get_or_create_account(db: Session, name: str, institution: str,
                           account_type: str, account_subtype: str = None) -> Account:
    acct = db.query(Account).filter(Account.name == name).first()
    if not acct:
        acct = Account(
            name=name, institution=institution,
            account_type=account_type, account_subtype=account_subtype,
        )
        db.add(acct)
        db.flush()
        print(f"  Created account: {name}")
    return acct


def get_category_id(db: Session, name: str) -> int | None:
    cat = db.query(Category).filter(Category.name == name).first()
    return cat.id if cat else None


CATEGORY_MAP = {
    "paycheck":        "Paycheck",
    "transfer":        "Uncategorized",
    "interest income": "Investment Return",
    "food & drink":    "Restaurants",
    "shopping":        "Online Shopping",
    "groceries":       "Groceries",
    "gas":             "Gas",
    "medical":         "Medical",
    "travel":          "Travel",
    "entertainment":   "Streaming",
    "personal care":   "Uncategorized",
    "home":            "Uncategorized",
    "insurance":       "Insurance",
    "utilities":       "Utilities",
    "automotive":      "Gas",
    "education":       "Education",
    "fees":            "Bank Fees",
    "gifts":           "Gifts",
}


def map_category(db: Session, raw: str) -> int | None:
    if not raw:
        return None
    lower = raw.lower().strip()
    for key, cat_name in CATEGORY_MAP.items():
        if key in lower:
            return get_category_id(db, cat_name)
    return get_category_id(db, "Uncategorized")


# ── Bank CSV importer (Checking / Savings / Credit Card) ─────────────────────

def import_bank_csv(db: Session, filepath: Path, acct_name: str,
                    institution: str, acct_type: str, acct_subtype: str = None):
    """
    Expected columns: Date, Description, Original Description, Category, Amount, Status
    Positive amount = credit/income, Negative = debit/expense in USAA format.
    We store positive = expense per app convention.
    """
    print(f"\nImporting {filepath.name} → '{acct_name}'")
    acct = get_or_create_account(db, acct_name, institution, acct_type, acct_subtype)

    df = pd.read_csv(filepath, dtype=str).fillna("")
    df.columns = df.columns.str.strip()
    imported = skipped = 0

    for _, row in df.iterrows():
        raw_date = row.get("Date", "").strip()
        dt = parse_date_utc(raw_date)
        if dt is None:
            skipped += 1
            continue

        # USAA CSV: positive=credit (money in), negative=debit (money out)
        raw_amt = to_float(row.get("Amount", "0"))
        # Convert to app convention: positive = expense
        amount = -raw_amt  # flip sign so outflows are positive

        desc = row.get("Description", row.get("Original Description", "")).strip()
        orig_desc = row.get("Original Description", "").strip()
        status = row.get("Status", "").strip().lower()
        pending = status == "pending"
        raw_cat = row.get("Category", "").strip()
        cat_id = map_category(db, raw_cat)

        # Deduplicate: same account + date + amount + description
        exists = db.query(Transaction).filter(
            Transaction.account_id == acct.id,
            Transaction.date == dt,
            Transaction.amount == amount,
            Transaction.description == desc,
        ).first()
        if exists:
            skipped += 1
            continue

        txn = Transaction(
            account_id=acct.id,
            category_id=cat_id,
            amount=amount,
            description=desc,
            merchant_name=desc[:60] if desc else None,
            date=dt,
            pending=pending,
            notes=orig_desc if orig_desc != desc else None,
            source="csv",
        )
        db.add(txn)
        imported += 1

    db.commit()
    print(f"  → Imported {imported} transactions, skipped {skipped}")
    return imported


# ── Fidelity ByBenefitType (XLSX) ─────────────────────────────────────────────

def import_fidelity_by_benefit(db: Session, filepath: Path):
    """
    Fidelity ByBenefitType xlsx has two sheets:
      - ESPP: INTC Employee Stock Purchase Plan purchases
      - Restricted Stock: INTC RSU grants with vest schedules
    """
    print(f"\nImporting {filepath.name} → Fidelity (ByBenefitType)")

    existing = db.query(InvestmentAccount).filter(
        InvestmentAccount.source_file == filepath.name,
        InvestmentAccount.institution == "Fidelity",
    ).first()
    if existing:
        print("  Already imported, skipping.")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        print(f"  ERROR reading xlsx: {e}")
        return

    acct = InvestmentAccount(
        name="Fidelity – Intel ESPP & Restricted Stock",
        institution="Fidelity",
        account_type="retirement",
        source_file=filepath.name,
        as_of_date=datetime.now(timezone.utc),
    )
    db.add(acct)
    db.flush()

    holding_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(v).strip() if v else "" for v in rows[0]]

        for row in rows[1:]:
            if not any(row):
                continue
            d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            record_type = str(d.get("Record Type", "") or "").strip()
            symbol = str(d.get("Symbol", "") or "").strip()

            if not record_type or not symbol:
                continue
            if record_type.lower() in ("totals", "vest schedule", "tax withholding"):
                continue

            # Determine shares: ESPP uses 'Net Shares', RSU uses 'Sellable Qty.'
            shares = (
                d.get("Net Shares") or d.get("Purchased Qty.") or
                d.get("Sellable Qty.") or d.get("Granted Qty.")
            )
            price = d.get("Purchase Price")

            extra = {k: str(v) for k, v in d.items() if v is not None and str(v).strip() not in ("", "None")}

            holding = InvestmentHolding(
                investment_account_id=acct.id,
                symbol=symbol,
                description=f"{sheet_name} – {record_type}",
                holding_type="stock",
                shares=float(shares) if shares is not None else None,
                price_per_share=float(price) if price is not None else None,
                market_value=None,  # no current market value in this file
                extra_json=json.dumps(extra),
                as_of_date=datetime.now(timezone.utc),
            )
            db.add(holding)
            holding_count += 1

    db.commit()
    print(f"  -> Created Fidelity ByBenefitType account id={acct.id}, {holding_count} holdings")


# ── Fidelity Total Value (XLSX) ────────────────────────────────────────────────

def import_fidelity_total_value(db: Session, filepath: Path):
    print(f"\nImporting {filepath.name} → Fidelity 401K total value snapshot")

    existing = db.query(InvestmentAccount).filter(
        InvestmentAccount.source_file == filepath.name,
    ).first()
    if existing:
        print("  Already imported, skipping.")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        print(f"  ERROR: {e}")
        return

    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if len(vals) < 2:
            continue
        # First non-None is name, second is total value (float)
        name = str(vals[0]).strip()
        total = float(vals[1]) if isinstance(vals[1], (int, float)) else to_float(str(vals[1]))
        if total == 0.0:
            continue
        # Extract account number from name like "*5105"
        m = re.search(r"\*(\d+)", name)
        acct_num = m.group(1)[-4:] if m else None

        acct = InvestmentAccount(
            name=name[:200],
            institution="Fidelity",
            account_type="retirement",
            account_number_last4=acct_num,
            total_value=total,
            source_file=filepath.name,
            as_of_date=datetime.now(timezone.utc),
        )
        db.add(acct)
        db.flush()
        print(f"  -> {name}: ${total:,.2f}")

    db.commit()


# ── Charles Schwab (XLSX) ──────────────────────────────────────────────────────

def _schwab_extract(cell_val: str, prefix: str) -> str:
    """Extract numeric string after a label prefix like 'Price$307.69' → '307.69'"""
    if not cell_val:
        return ""
    s = str(cell_val).replace("\xa0", " ")
    # Remove the prefix label then strip non-numeric prefix chars
    if prefix.lower() in s.lower():
        s = re.sub(re.escape(prefix), "", s, flags=re.IGNORECASE).strip()
    return s


def import_schwab(db: Session, filepath: Path):
    """
    Schwab xlsx export stores each cell as 'LabelValue' concatenated strings.
    E.g. col[1]='Quantity120', col[2]='Price$307.69', col[4]='Market Value$36,922.80'
    We parse by regex on each cell value.
    """
    print(f"\nImporting {filepath.name} → Charles Schwab")

    existing = db.query(InvestmentAccount).filter(
        InvestmentAccount.source_file == filepath.name,
    ).first()
    if existing:
        print("  Already imported, skipping.")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        print(f"  ERROR: {e}")
        return

    all_rows = list(ws.iter_rows(values_only=True))

    # Find total from "Cash & Money Market Total" row or sum of group totals
    total_val = 0.0
    group_totals = []
    for row in all_rows:
        cell0 = str(row[0] or "")
        if "Total" in cell0 and "security group" not in cell0 and "Cash &" not in cell0:
            # e.g. 'Equities Total', 'Mutual Funds Total', 'Cash & Money Market Total'
            mkt_cell = str(row[4] or "") if len(row) > 4 else ""
            m = re.search(r"\$([\d,]+\.?\d*)", mkt_cell)
            if m:
                group_totals.append(to_float(m.group(1)))
    total_val = sum(group_totals)

    acct = InvestmentAccount(
        name="Charles Schwab Individual Brokerage",
        institution="Charles Schwab",
        account_type="brokerage",
        total_value=total_val or None,
        source_file=filepath.name,
        as_of_date=datetime.now(timezone.utc),
    )
    db.add(acct)
    db.flush()

    holding_count = 0
    current_type = "stock"
    for row in all_rows:
        cell0 = str(row[0] or "").strip().replace("\xa0", " ")
        if not cell0:
            continue

        # Detect section header
        if cell0 in ("Equities",):
            current_type = "stock"
            continue
        if cell0 in ("Mutual Funds",):
            current_type = "mutual_fund"
            continue
        if cell0 in ("Cash & Money Market",):
            current_type = "cash"
            continue

        # Skip non-data rows
        if any(x in cell0 for x in ("Symbol", "Positions", "Click to", "Total", "*.  Data", "*.")):
            continue
        if cell0.startswith("Cash &") and "Total" in cell0:
            continue

        # Is this a holdings row? Check cell[1] for Quantity pattern
        qty_cell  = str(row[1] or "") if len(row) > 1 else ""
        if not re.search(r"Quantity[\d.]+", qty_cell):
            continue

        symbol = cell0
        qty_m     = re.search(r"Quantity([\d.]+)", qty_cell)
        price_m   = re.search(r"Price\$([\d,]+\.?\d*)", str(row[2] or ""))
        mktval_m  = re.search(r"Market Value\$([\d,]+\.?\d*)", str(row[4] or ""))
        cost_m    = re.search(r"Cost Basis\$([\d,]+\.?\d*)", str(row[6] or ""))
        gain_m    = re.search(r"Gain Loss([+-]?\$[\d,]+\.?\d*)", str(row[7] or ""))

        # Description is usually on the NEXT row (col 0), no Quantity
        desc = symbol  # will be overridden below if next row is a name row

        holding = InvestmentHolding(
            investment_account_id=acct.id,
            symbol=symbol,
            description=desc,
            holding_type=current_type,
            shares=to_float(qty_m.group(1)) if qty_m else None,
            price_per_share=to_float(price_m.group(1)) if price_m else None,
            market_value=to_float(mktval_m.group(1)) if mktval_m else None,
            cost_basis=to_float(cost_m.group(1)) if cost_m else None,
            unrealized_gain_loss=to_float(gain_m.group(1).replace("$", "")) if gain_m else None,
            as_of_date=datetime.now(timezone.utc),
        )
        db.add(holding)
        holding_count += 1

    db.commit()
    print(f"  -> Created Schwab account id={acct.id}, total=${total_val:,.2f}, {holding_count} holdings")


# ── Edward Jones (CSV) ────────────────────────────────────────────────────────

def import_edward_jones(db: Session, filepath: Path):
    print(f"\nImporting {filepath.name} → Edward Jones")

    existing = db.query(InvestmentAccount).filter(
        InvestmentAccount.source_file == filepath.name,
    ).first()
    if existing:
        print("  Already imported, skipping.")
        return

    # Row 1 is "Joint-1 Wednesday…" — skip 3 rows to get real header
    df = pd.read_csv(filepath, dtype=str, skiprows=3).fillna("")
    df.columns = df.columns.str.strip()

    total_val = 0.0
    for _, row in df.iterrows():
        sym = str(row.get("SYMBOL/CUSIP", "")).strip()
        row_type = str(row.get("TYPE", "")).strip()
        val = to_float(row.get("ESTIMATED CURRENT VALUE", "0"))
        # Last row has no TYPE and no symbol — it's the grand total
        if not sym and not row_type and val > 0:
            total_val = val
            break

    acct_nick = df["ACCOUNT NICKNAME"].dropna().iloc[0] if "ACCOUNT NICKNAME" in df.columns else "Edward Jones"
    acct = InvestmentAccount(
        name=f"Edward Jones – {acct_nick}"[:200],
        institution="Edward Jones",
        account_type="retirement",
        total_value=total_val or None,
        source_file=filepath.name,
        as_of_date=datetime.now(timezone.utc),
    )
    db.add(acct)
    db.flush()

    TYPE_MAP = {
        "stock": "stock", "mutual fund": "mutual_fund", "etf": "etf",
        "cash": "cash", "insured bank deposit": "cash", "bond": "bond",
        "cd": "cd",
    }

    for _, row in df.iterrows():
        symbol = str(row.get("SYMBOL/CUSIP", "")).strip()
        desc   = str(row.get("DESCRIPTION", "")).strip()
        raw_type = str(row.get("TYPE", "")).strip().lower()
        htype = TYPE_MAP.get(raw_type, "other")
        # Skip the grand total summary row
        if not symbol and not desc:
            continue
        # Skip blank type rows (total row)
        if not raw_type and not symbol:
            continue

        ann_rate_str = str(row.get("ANNUALIZED RATE OF RETURN", "")).strip()
        ann_rate = to_float(ann_rate_str) if ann_rate_str else None

        holding = InvestmentHolding(
            investment_account_id=acct.id,
            symbol=symbol or None,
            description=desc or None,
            holding_type=htype,
            shares=to_float(row.get("SHARES", "0")) or None,
            price_per_share=to_float(row.get("PRICE/SHARE", "0")) or None,
            market_value=to_float(row.get("ESTIMATED CURRENT VALUE", "0")) or None,
            cost_basis=to_float(row.get("COST BASIS", "0")) or None,
            unrealized_gain_loss=to_float(row.get("UNREALIZED GAIN LOSS", "0")) or None,
            realized_gain_loss=to_float(row.get("REALIZED GAIN LOSS", "0")) or None,
            annualized_return_pct=ann_rate,
            return_dollars=to_float(row.get("RETURN IN DOLLARS", "0")) or None,
            as_of_date=datetime.now(timezone.utc),
        )
        db.add(holding)

    db.commit()
    print(f"  → Created Edward Jones account id={acct.id}, total=${total_val:,.2f}")


# ── E*Trade (CSV) ─────────────────────────────────────────────────────────────

def import_etrade(db: Session, filepath: Path):
    print(f"\nImporting {filepath.name} → E*Trade")

    existing = db.query(InvestmentAccount).filter(
        InvestmentAccount.source_file == filepath.name,
    ).first()
    if existing:
        print("  Already imported, skipping.")
        return

    # E*Trade CSV has mixed section headers and varying column counts — read raw lines
    text = filepath.read_text(encoding="utf-8-sig", errors="replace")
    lines = text.splitlines()

    total_val = 0.0
    acct_num  = None
    cash_val  = 0.0
    as_of_str = None

    for line in lines:
        # Account number like "Individual Brokerage -1358"
        m = re.search(r"Individual Brokerage -(\d+)", line)
        if m:
            acct_num = m.group(1)[-4:]

        # Net Account Value row: "Individual Brokerage -1358",5622.00,...
        if acct_num and acct_num in line:
            parts = [p.strip().strip('"') for p in line.split(",")]
            for p in parts:
                f = to_float(p)
                if f > 0:
                    total_val = max(total_val, f)

        # CASH row
        if line.startswith("CASH,"):
            parts = [p.strip().strip('"') for p in line.split(",")]
            # Last numeric value is the cash balance
            nums = [to_float(p) for p in parts if to_float(p) > 0]
            if nums:
                cash_val = max(nums)

        # Generated timestamp
        if line.startswith("Generated at"):
            as_of_str = line.replace("Generated at", "").strip()

    as_of_dt = parse_date_utc(as_of_str) if as_of_str else datetime.now(timezone.utc)

    acct = InvestmentAccount(
        name=f"E*Trade Individual Brokerage{' *'+acct_num if acct_num else ''}",
        institution="E*Trade",
        account_type="brokerage",
        account_number_last4=acct_num,
        total_value=total_val or cash_val or None,
        source_file=filepath.name,
        as_of_date=as_of_dt,
    )
    db.add(acct)
    db.flush()

    bal = total_val or cash_val
    if bal > 0:
        holding = InvestmentHolding(
            investment_account_id=acct.id,
            symbol="CASH",
            description="Cash Balance",
            holding_type="cash",
            market_value=bal,
            as_of_date=as_of_dt,
        )
        db.add(holding)

    db.commit()
    print(f"  -> Created E*Trade account id={acct.id}, total=${bal:,.2f}")


# ── Irregular Expense History (CSV) ───────────────────────────────────────────

def import_irregular_expenses(db: Session, filepath: Path):
    print(f"\nImporting {filepath.name} → irregular_expenses")

    existing_count = db.query(IrregularExpense).count()
    if existing_count > 0:
        print(f"  Already have {existing_count} rows — skipping.")
        return

    df = pd.read_csv(filepath, dtype=str).fillna("")
    df.columns = df.columns.str.strip()

    COL_MAP = {
        "Date":                  "date",
        "Quarter":               "quarter",
        "Month":                 "month",
        "Month Number":          "month_number",
        "Work Week":             "work_week",
        "Week of the Month":     "week_of_month",
        "Day":                   "day_of_week",
        "Groceries":             "groceries",
        "Toilettries":           "toiletries",
        "Dining Out":            "dining_out",
        "Alcohol":               "alcohol",
        "Children":              "children",
        "Car Maintenance":       "car_maintenance",
        "Fuel":                  "fuel",
        "Home Maint./Projects":  "home_maintenance",
        "Clothing":              "clothing",
        "Health":                "health",
        "Fun ":                  "fun",
        "Gifts":                 "gifts",
        "Party":                 "party",
        "Miscellaneous ":        "miscellaneous",
        "Travel":                "travel",
        "Total":                 "total",
        "Notes":                 "notes",
    }

    # Pre-aggregate duplicate dates (same date appears with different format)
    from collections import defaultdict
    agg: dict = defaultdict(lambda: {
        "quarter": None, "month": None, "month_number": None,
        "work_week": None, "week_of_month": None, "day_of_week": None,
        "groceries": 0.0, "toiletries": 0.0, "dining_out": 0.0,
        "alcohol": 0.0, "children": 0.0, "car_maintenance": 0.0,
        "fuel": 0.0, "home_maintenance": 0.0, "clothing": 0.0,
        "health": 0.0, "fun": 0.0, "gifts": 0.0, "party": 0.0,
        "miscellaneous": 0.0, "travel": 0.0, "total": 0.0, "notes": None,
    })

    skipped = 0
    for _, row in df.iterrows():
        raw_date = row.get("Date", "").strip()
        d = parse_date_only(raw_date)
        if d is None:
            skipped += 1
            continue

        def gf(col): return to_float(row.get(col, "0"))
        def gi(col):
            v = row.get(col, "")
            try:
                return int(float(v)) if str(v).strip() else None
            except Exception:
                return None

        a = agg[d]
        a["quarter"]      = a["quarter"]      or gi("Quarter")
        a["month"]        = a["month"]        or row.get("Month", "").strip() or None
        a["month_number"] = a["month_number"] or gi("Month Number")
        a["work_week"]    = a["work_week"]    or gi("Work Week")
        a["week_of_month"]= a["week_of_month"]or gi("Week of the Month")
        a["day_of_week"]  = a["day_of_week"]  or row.get("Day", "").strip() or None
        a["groceries"]    += gf("Groceries")
        a["toiletries"]   += gf("Toilettries")
        a["dining_out"]   += gf("Dining Out")
        a["alcohol"]      += gf("Alcohol")
        a["children"]     += gf("Children")
        a["car_maintenance"] += gf("Car Maintenance")
        a["fuel"]         += gf("Fuel")
        a["home_maintenance"] += gf("Home Maint./Projects")
        a["clothing"]     += gf("Clothing")
        a["health"]       += gf("Health")
        a["fun"]          += gf("Fun ")
        a["gifts"]        += gf("Gifts")
        a["party"]        += gf("Party")
        a["miscellaneous"]  += gf("Miscellaneous ")
        a["travel"]       += gf("Travel")
        a["total"]        += gf("Total")
        note = row.get("Notes", "").strip()
        if note:
            a["notes"] = (a["notes"] + "; " + note) if a["notes"] else note

    imported = 0
    for d, a in sorted(agg.items()):
        expense = IrregularExpense(date=d, **a)
        db.add(expense)
        imported += 1
        if imported % 500 == 0:
            db.commit()
            print(f"  ... {imported} rows committed")

    db.commit()
    print(f"  -> Imported {imported} daily expense rows, skipped {skipped} unparseable")


# ── Regular Expense History (CSV) ─────────────────────────────────────────────

def import_regular_expenses(db: Session, filepath: Path):
    print(f"\nImporting {filepath.name} → regular_expenses")

    existing_count = db.query(RegularExpense).count()
    if existing_count > 0:
        print(f"  Already have {existing_count} rows — skipping.")
        return

    df = pd.read_csv(filepath, dtype=str).fillna("")
    df.columns = df.columns.str.strip()

    imported = skipped = 0
    for _, row in df.iterrows():
        month_name = row.get("Month", "").strip()
        year_str   = row.get("Year", "").strip()
        month_num  = row.get("Month Number", "").strip()
        if not month_name or not year_str:
            skipped += 1
            continue
        try:
            year = int(float(year_str))
            mnum = int(float(month_num)) if month_num else 0
        except (ValueError, TypeError):
            skipped += 1
            continue

        def gf(col): return to_float(row.get(col, "0"))

        reg = RegularExpense(
            month=month_name,
            month_number=mnum,
            year=year,
            electricity=gf("Electricity"),
            water=gf("Water"),
            garbage=gf("Garbage"),
            internet=gf("Internet"),
            cell_phone=gf("Cell Phone"),
            insurance=gf("Insurance"),
            loans_credit=gf("Loans/Credit"),
            total=gf("Total"),
            four_month_average=gf("4 Month Average") or None,
        )
        db.add(reg)
        imported += 1

    db.commit()
    print(f"  → Imported {imported} monthly expense rows, skipped {skipped}")


# ── Reconciliation pass ────────────────────────────────────────────────────────

def run_reconciliation(db: Session):
    """
    For irregular & regular expense rows from 2024-09-18 onward,
    compare their monthly totals to actual bank transactions.
    Writes reconcile_delta + reconcile_notes back to each row.
    """
    from datetime import date as date_type
    from sqlalchemy import func

    print("\nRunning reconciliation (2024-09-18 → today)…")
    start = date_type(2024, 9, 18)
    today = date_type.today()

    # Sum bank transactions (expenses) by year+month
    from datetime import datetime as dt_type
    txn_q = db.query(
        func.strftime("%Y", Transaction.date).label("yr"),
        func.strftime("%m", Transaction.date).label("mo"),
        func.sum(Transaction.amount).label("total"),
    ).filter(
        Transaction.date >= dt_type(2024, 9, 18, tzinfo=timezone.utc),
        Transaction.amount > 0,
        Transaction.source.in_(["csv", "manual", "plaid"]),
    ).group_by("yr", "mo").all()
    txn_totals = {(r.yr, r.mo): r.total or 0 for r in txn_q}

    # ── Reconcile irregular expenses ──
    irr_rows = db.query(IrregularExpense).filter(
        IrregularExpense.date >= start,
        IrregularExpense.date <= today,
    ).all()

    irr_monthly: dict[tuple, list] = {}
    for r in irr_rows:
        key = (str(r.date.year), f"{r.date.month:02d}")
        irr_monthly.setdefault(key, []).append(r)

    for key, rows in irr_monthly.items():
        hist_total = sum(r.total for r in rows)
        txn_total  = txn_totals.get(key, 0)
        delta = txn_total - hist_total
        note  = f"Bank txns ${txn_total:.2f} vs history ${hist_total:.2f}"
        for r in rows:
            r.reconcile_delta = round(delta, 2)
            r.reconcile_notes = note
            r.reconciled      = abs(delta) < 100

    # ── Reconcile regular expenses ──
    reg_rows = db.query(RegularExpense).filter(
        (RegularExpense.year > 2024) |
        ((RegularExpense.year == 2024) & (RegularExpense.month_number >= 9))
    ).all()

    for r in reg_rows:
        key = (str(r.year), f"{r.month_number:02d}")
        txn_total = txn_totals.get(key, 0)
        delta = txn_total - r.total
        r.reconcile_delta = round(delta, 2)
        r.reconcile_notes = f"Bank txns ${txn_total:.2f} vs history ${r.total:.2f}"
        r.reconciled      = abs(delta) < 100

    db.commit()
    print(f"  → Reconciled {len(irr_rows)} irregular rows and {len(reg_rows)} regular rows.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    db: Session = SessionLocal()
    try:
        print("=" * 60)
        print("Finance App — Comprehensive Data Seed")
        print("=" * 60)

        # Bank accounts
        import_bank_csv(db, DATA_DIR / "Checking.csv",
                        "USAA Checking", "USAA", "depository", "checking")
        import_bank_csv(db, DATA_DIR / "Savings.csv",
                        "USAA Savings", "USAA", "depository", "savings")
        import_bank_csv(db, DATA_DIR / "USAA CreditCard.csv",
                        "USAA Credit Card", "USAA", "credit", "credit card")

        # Investment / retirement accounts
        import_fidelity_by_benefit(db, DATA_DIR / "ByBenefitType_expanded.xlsx")
        import_fidelity_total_value(db, DATA_DIR / "FidelityTotalValue.xlsx")
        import_schwab(db,          DATA_DIR / "charlesschwab copy.xlsx")

        ej_files = list(DATA_DIR.glob("Edward Jones holdings*.csv"))
        for f in ej_files:
            import_edward_jones(db, f)

        etrade_files = list({f.name: f for f in list(DATA_DIR.glob("ETrade*.csv")) + list(DATA_DIR.glob("eTrade*.csv"))}.values())
        for f in etrade_files:
            import_etrade(db, f)

        # Expense history
        import_irregular_expenses(db, DATA_DIR / "Irregular Expense History.csv")
        import_regular_expenses(db,   DATA_DIR / "Regular Expense History.csv")

        # Reconciliation
        run_reconciliation(db)

        print("\n" + "=" * 60)
        print("Seed complete!")

        # Summary
        print(f"\n  Accounts (bank):       {db.query(Account).count()}")
        print(f"  Transactions:          {db.query(Transaction).count()}")
        print(f"  Investment accounts:   {db.query(InvestmentAccount).count()}")
        print(f"  Investment holdings:   {db.query(InvestmentHolding).count()}")
        print(f"  Irregular exp rows:    {db.query(IrregularExpense).count()}")
        print(f"  Regular exp rows:      {db.query(RegularExpense).count()}")
        print("=" * 60)

    except Exception as e:
        db.rollback()
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
